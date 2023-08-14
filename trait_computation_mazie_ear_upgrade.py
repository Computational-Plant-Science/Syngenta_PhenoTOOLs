'''
Name: trait_extract_parallel.py

Version: 1.0

Summary: Extract maize ear traits 
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2022-09-29

USAGE:

time python3 trait_computation_mazie_ear_upgrade.py -p ~/example/plant_test/seeds/test/ -ft png -s Lab -c 0 -ne 5 -min 250000


'''

# import necessary packages
import os
import glob
import utils

from collections import Counter

import argparse

from sklearn.cluster import KMeans
from sklearn.cluster import MiniBatchKMeans

from skimage.feature import peak_local_max
from skimage.morphology import medial_axis
from skimage import img_as_float, img_as_ubyte, img_as_bool, img_as_int
from skimage import measure
from skimage.color import rgb2lab, deltaE_cie76
from skimage import morphology
from skimage.segmentation import clear_border, watershed
from skimage.measure import regionprops

from scipy.spatial import distance as dist
from scipy import optimize
from scipy import ndimage 
from scipy.interpolate import interp1d
from scipy.spatial.distance import pdist

from skan.csr import skeleton_to_csgraph
from skan import Skeleton, summarize, draw


import imutils
from imutils import perspective

import numpy as np
import argparse
import cv2

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import collections

import math
import openpyxl
import csv
    
from tabulate import tabulate

from pathlib import Path 
from pylibdmtx.pylibdmtx import decode
import re


import psutil
import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from rembg import remove

import natsort 

import warnings
warnings.filterwarnings("ignore")


MBFACTOR = float(1<<20)



def mkdir(path):
    
    """create folder and path to store the output results
    
    Inputs: 
    
        path: result path
        
       
    Returns:
    
        create path and folder if not exist  
        
    """   
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False
        


def adaptive_threshold(masked_image, GaussianBlur_ksize, blockSize, weighted_mean):
    
    """compute thresh image using adaptive threshold Method
    
    Inputs: 
    
        maksed_img: masked image contains only target objects
        
        GaussianBlur_ksize: Gaussian Kernel Size 
        
        blockSize: size of the pixelneighborhood used to calculate the threshold value
        
        weighted_mean: the constant used in the both methods (subtracted from the mean or weighted mean).

    Returns:
        
        thresh_adaptive_threshold: thresh image using adaptive thrshold Method
        
        maksed_img_adaptive_threshold: masked image using thresh_adaptive_threshold

    """
    ori = masked_image.copy()
    
    if len(ori.shape)> 2:
        # convert the image to grayscale and blur it slightly
        gray = cv2.cvtColor(masked_image, cv2.COLOR_BGR2GRAY)
    else:
        gray = ori
    
    # blurring it . Applying Gaussian blurring with a GaussianBlur_ksize×GaussianBlur_ksize kernel 
    # helps remove some of the high frequency edges in the image that we are not concerned with and allow us to obtain a more “clean” segmentation.
    blurred = cv2.GaussianBlur(gray, (GaussianBlur_ksize, GaussianBlur_ksize), 0)

    # adaptive method to be used. 'ADAPTIVE_THRESH_MEAN_C' or 'ADAPTIVE_THRESH_GAUSSIAN_C'
    thresh_adaptive_threshold = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, blockSize, weighted_mean)

    # apply individual object mask
    maksed_img_adaptive_threshold = cv2.bitwise_and(ori, ori.copy(), mask = ~thresh_adaptive_threshold)

    return thresh_adaptive_threshold, maksed_img_adaptive_threshold




def sort_contours(cnts, method = "left-to-right"):
    
    """sort contours based on user defined method
    
    Inputs: 
    
        cnts: contours extracted from mask image
        
        method: user defined method, default was "left-to-right"
        

    Returns:
    
        sorted_cnts: list of sorted contours 
        
    """   
    
    
    # initialize the reverse flag and sort index
    reverse = False
    i = 0
    
    # handle if we need to sort in reverse
    if method == "right-to-left" or method == "bottom-to-top":
        reverse = True
        
    # handle if we are sorting against the y-coordinate rather than
    # the x-coordinate of the bounding box
    if method == "top-to-bottom" or method == "bottom-to-top":
        i = 1
        
    # construct the list of bounding boxes and sort them from top to bottom
    boundingBoxes = [cv2.boundingRect(c) for c in cnts]
    
    (sorted_cnts, boundingBoxes) = zip(*sorted(zip(cnts, boundingBoxes), key=lambda b:b[1][i], reverse=reverse))
    
    # return the list of sorted contours and bounding boxes
    return sorted_cnts



def mutilple_objects_detection(orig):
    """segment mutiple objects in image, for maize ear imagem, based on the protocal, number of ears ranges form 1 to 5.
    
    Inputs: 
    
        orig: image of plant object
        
    Returns:
    
        images: each image contains one maize ear on the left/right side 

    """   
    # remove the back ground using Rembg is a tool to remove images background.
    orig_mask = remove(orig.copy(), only_mask=True)
    
    orig_forground = remove(orig.copy())
    
    # get the dimension of the image
    #height, width, channels = orig_foreground.shape
    
    
    return orig_mask, orig_forground
    


# segment mutiple objects in image, for maize ear image, based on the protocal, shoudl be two objects. 
def mutilple_objects_seg(orig, channel, size_kernel):

    """segment mutiple objects in image, for maize ear image, based on the protocal, should be less than 5 objects.
    
    Inputs: 
    
        orig: image of plant object
        
    Returns:
    
        left_img, right_img: left/right image contains each maize ear on the left/right side 
        
        mask_seg_gray: 
        
        img_overlay:
        
        cnt_area: 
        
    """   
    # apply smooth filtering of the image at the color level.
    shifted = cv2.pyrMeanShiftFiltering(orig, 21, 70)
    
    #shifted = cv2.pyrMeanShiftFiltering(orig, 21, 10)

    # get the dimension of the image
    height, width, channels = orig.shape
    
    '''
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)

    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)

    else:
        colorSpace = 'bgr'  # set for file naming purposes
    
        

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
    '''

    # Convert mean shift image from BRG color space to LAB space and extract B channel
    L, A, B = cv2.split(cv2.cvtColor(shifted, cv2.COLOR_BGR2LAB))
    
    # convert the mean shift image to grayscale, then apply Otsu's thresholding
    #gray = cv2.cvtColor(shifted, cv2.COLOR_BGR2GRAY)
    
    if channel == 'B':
        
        thresh = cv2.threshold(B, 128, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    elif channel == 'A':
        
        thresh = cv2.threshold(A, 128, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    elif channel == 'L':
        
        thresh = cv2.threshold(L, 128, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    # Taking a matrix of size_ker as the kernel
    
    #size_kernel = 10
    
    kernel = np.ones((size_kernel, size_kernel), np.uint8)
    
    # apply morphological operations to remove noise
    thresh_dilation = cv2.dilate(thresh, kernel, iterations=1)
    
    thresh_erosion = cv2.erode(thresh, kernel, iterations=1)
    

    
    # find contours in the thresholded image
    cnts = cv2.findContours(thresh_erosion.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    cnts = imutils.grab_contours(cnts)
    
    # sort the contour based on area size from largest to smallest, and get the first two max contours
    cnts_sorted = sorted(cnts, key = cv2.contourArea, reverse = True)[0:n_ear]

    # sort the contours from left to right
    cnts_sorted = sort_contours(cnts_sorted, method = "left-to-right")
    
    #print("cv2.contourArea(cnts_sorted[0]), cv2.contourArea(cnts_sorted[1])")
    #print(cv2.contourArea(cnts_sorted[0]), cv2.contourArea(cnts_sorted[1]))
    
    #print("left-to-right")
    #print(len(cnts_sorted))
    

    # initialize variables to record the centers, area of contours
    center_locX = []
    center_locY = []
    cnt_area = [0] * n_ear
    
    # initialize empty mask image
    img_thresh = np.zeros(orig.shape, np.uint8)
    
    # initialize background image to draw the contours
    img_overlay_bk = orig
    
    # loop over the selected contours
    for idx, c in enumerate(cnts_sorted):
        
        # compute the center of the contour
        M = cv2.moments(c)
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        
        # record the center coordinates
        center_locX.append(cX)
        center_locY.append(cY)

        # get the contour area
        cnt_area[idx] = cv2.contourArea(c)
        
        # draw the contour and center of the shape on the image
        #img_overlay = cv2.drawContours(img_overlay_bk, [c], -1, (0, 255, 0), 2)
        mask_seg = cv2.drawContours(img_thresh, [c], -1, (255,255,255), -1)
        #center_result = cv2.circle(img_thresh, (cX, cY), 14, (0, 0, 255), -1)
        #img_overlay = cv2.putText(img_overlay_bk, "{}".format(idx +1), (cX - 20, cY - 20), cv2.FONT_HERSHEY_SIMPLEX, 5.5, (255, 0, 0), 5)
        
    
    # get the middle point coordinate of the two centers of the contours
    #divide_X = int(sum(center_locX) / len(center_locX))
    #divide_Y = int(sum(center_locY) / len(center_locY))
    
    # get the left and right segmentation of the image 
    #left_img = orig[0:height, 0:divide_X]
    #right_img = orig[0:height, divide_X:width]
    

    # convert the mask image to gray format
    mask_seg_gray = cv2.cvtColor(mask_seg, cv2.COLOR_BGR2GRAY)
    
    
    #return left_img, right_img, mask_seg_gray, img_overlay, cnt_area
    
    return mask_seg_gray
    
    
    

# color clustering based object segmentation
def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):
    
    """color clustering based object segmentation
    
    Inputs: 
    
        image: image contains the plant objects
        
        args_colorspace: user-defined color space for clustering 
        
        args_channels: user-defined color channel for clustering 
        
        args_num_clusters: number of clustering
        
    Returns:
    
        img_thresh: mask image with the segmentation of the plant object 
        
    """   
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
        
    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        
    else:
        colorSpace = 'bgr'  # set for file naming purposes
        
        
    #image = cv2.pyrMeanShiftFiltering(image, 21, 70)

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
    
    # get the dimension of image 
    (width, height, n_channel) = image.shape

    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    
    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    #define number of cluster
    numClusters = max(2, args_num_clusters)
    
    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i
    
    ret, thresh = cv2.threshold(kmeansImage,0,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    

    # clean the border of mask image
    if np.count_nonzero(thresh) > 0:
        
        thresh_cleaned = clear_border(thresh)
    else:
        thresh_cleaned = thresh
    
    # get the connected Components in the mask image
    nb_components, output, stats, centroids = cv2.connectedComponentsWithStats(thresh_cleaned, connectivity = 8)

    # stats[0], centroids[0] are for the background label. ignore
    # cv2.CC_STAT_LEFT, cv2.CC_STAT_TOP, cv2.CC_STAT_WIDTH, cv2.CC_STAT_HEIGHT
    
    # get all connected Components's area value
    sizes = stats[1:, cv2.CC_STAT_AREA]

    # remove background component
    nb_components = nb_components - 1
    
    # create an empty mask image and fill the detected connected components
    img_thresh = np.zeros([width, height], dtype=np.uint8)
    
    #for every component in the image, keep it only if it's above min_size
    for i in range(0, nb_components):
        
        if (sizes[i] >= min_size):
        
            img_thresh[output == i + 1] = 255
        
    
    #if mask contains mutiple non-conected parts, combine them into one. 
    contours, hier = cv2.findContours(img_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if len(contours) > 1:
        
        print("mask contains mutiple non-connected parts, combine them into one\n")
        
        # create an size 10 kernel
        kernel = np.ones((10,10), np.uint8)
        
        # image dilation
        dilation = cv2.dilate(img_thresh.copy(), kernel, iterations = 1)
        
        # image closing
        closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
        
        # use the final closing result as mask
        img_thresh = closing


    return img_thresh
    




def percentage(part, whole):

    """compute percentage value
    
    Inputs: 
    
        part, whole: the part and whole value
        
       
    Returns:
    
        string type of the percentage in decimals 
        
    """   
    #percentage = "{:.0%}".format(float(part)/float(whole))

    percentage = "{:.2f}".format(float(part)/float(whole))

    return str(percentage)



 
def midpoint(ptA, ptB):

    """compute middle point of two points in 2D coordinates
    
    Inputs: 
    
        ptA, ptB: coordinates of two points
        
    Returns:
    
        coordinates of the middle point
        
    """   
    
    return ((ptA[0] + ptB[0]) * 0.5, (ptA[1] + ptB[1]) * 0.5)




'''
def adaptive_threshold_external(img):
    
    """compute thresh image using adaptive threshold Method
    
    Inputs: 
    
        img: image data

    Returns:
        
        mask_external: segmentation mask for external contours
        
        trait_img: original image overlay with bounding rect and contours

    """
    
    # obtain image dimension
    img_height, img_width, n_channels = img.shape
    
    orig = img.copy()
    
    # convert the image to grayscale and blur it slightly
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # set the parameters for adoptive threshholding method
    GaussianBlur_ksize = 5
    blockSize = 41
    weighted_mean = 10

    # adoptive threshholding method to the masked image from mutilple_objects_seg
    #(thresh_adaptive_threshold, maksed_img_adaptive_threshold) = adaptive_threshold(gray, GaussianBlur_ksize, blockSize, weighted_mean)
    

    
    # blurring it . Applying Gaussian blurring with a GaussianBlur_ksize×GaussianBlur_ksize kernel 
    # helps remove some of the high frequency edges in the image that we are not concerned with and allow us to obtain a more “clean” segmentation.
    blurred = cv2.GaussianBlur(gray, (GaussianBlur_ksize, GaussianBlur_ksize), 0)

    # adaptive method to be used. 'ADAPTIVE_THRESH_MEAN_C' or 'ADAPTIVE_THRESH_GAUSSIAN_C'
    thresh_adaptive_threshold = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 41, 10)

    # apply individual object mask
    maksed_img_adaptive_threshold = cv2.bitwise_and(orig, orig.copy(), mask = ~thresh_adaptive_threshold)
    
    
    #find contours and get the external one
    contours, hier = cv2.findContours(thresh_adaptive_threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # sort the contours based on area from largest to smallest
    contours_sorted = sorted(contours, key = cv2.contourArea, reverse = True)
    
    #contours_sorted = contours
    
    #select correct contours 
    ##########################################################################
    rect_area_rec = []
    
    # save all the boundingRect area for each contour 
    for index, c in enumerate(contours_sorted):
        
        #get the bounding rect
            (x, y, w, h) = cv2.boundingRect(c)
            
            rect_area_rec.append(w*h)
    
    # sort all contours according to the boundingRect area size in descending order
    idx_sort = [i[0] for i in sorted(enumerate(rect_area_rec), key=lambda k: k[1], reverse=True)]
    
    
    # initialize parametrs for first 3 biggest boundingRect
    rect_center_rec = []
    rect_size_rec = []
    
    # loop to record the center and size of the three boundingRect
    for index, value in enumerate(idx_sort[0:3]):
        
        # get the contour by index
        c = contours_sorted[value]
        
        #get the bounding rect
        (x, y, w, h) = cv2.boundingRect(c)
        
        center = (x, y)
        rect_center_rec.append(center)
        rect_size_rec.append(w*h)

    
    # extarct x value from center coordinates
    x_center = [i[0] for i in rect_center_rec]
    

    #######################################################################################3
    # choose the adjacent center pair among all three centers 
    if ((abs(x_center[0] - x_center[2]) < abs(x_center[0] - x_center[1])) or (abs(x_center[1] - x_center[2]) < abs(x_center[0] - x_center[2]))) \
    and ((abs(x_center[0] - x_center[1]) > abs(x_center[0] - x_center[2])) or (abs(x_center[0] - x_center[1]) > abs(x_center[1] - x_center[2]))):
            print("select objects successful...")
    
    else:
        
        # compute the average distance between adjacent center pair 
        avg_dist = sum(pdist(rect_center_rec))/len(pdist(rect_center_rec))
        
        # get the index of the min distance 
        idx_min = [i for i, j in enumerate(pdist(rect_center_rec)) if j < avg_dist]
        
        # choose the potiential candidate from the adjacent pair
        rect_size_rec_sel = rect_size_rec[idx_min[0]: int(idx_min[0]+2)]
        
        # get the index of the false contour
        idx_delete = np.argmin(rect_size_rec_sel)
        
        # delete the index of the false contour
        idx_sort.pop(idx_delete)
        
    
    ####################################################################################3
    area_rec = []
    
    trait_img = orig
    
    mask = np.zeros(gray.shape, dtype = "uint8")
    
    
    
    for index, value in enumerate(idx_sort):
        
        if index < 2:
             
            # visualize only the two external contours and its bounding box
            c = contours_sorted[value]
            
            # compute the convex hull of the contour
            hull = cv2.convexHull(c)
            
            # compute the area of the convex hull 
            hullArea = float(cv2.contourArea(hull))
            
            # save the convex hull area
            area_rec.append(hullArea)
            
            #get the bounding rect
            (x, y, w, h) = cv2.boundingRect(c)
            
            # draw a rectangle to visualize the bounding rect
            #trait_img = cv2.drawContours(orig, c, -1, (255, 255, 0), 3)
            
            #area_c_cmax = cv2.contourArea(c)
            
            trait_img = cv2.putText(orig, "#{0}".format(index), (int(x) - 10, int(y) - 20),cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 255), 2)
            
            # draw a green rectangle to visualize the bounding rect
            trait_img = cv2.rectangle(orig, (x, y), (x+w, y+h), (255, 255, 0), 4)
            
            # draw convexhull in red color
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255), 4)
            
            mask_external = cv2.drawContours(mask, [hull], -1, (255, 255, 255), -1)
            
    # compute the average area of the ear objects
    #external_contour_area = sum(area_rec)/len(area_rec)
    

 
    #define result path for labeled images
    #result_img_path = save_path + str(filename[0:-4]) + '_ctr.png'
    
    # save results
    #cv2.imwrite(result_img_path, trait_img)
    
    
    #define result path for labeled images
    #result_img_path = save_path + str(filename[0:-4]) + '_mask_external.png'
    
    # save results
    #cv2.imwrite(result_img_path, mask_external)
    
    return mask_external, trait_img
    
'''





# convert RGB value to HEX format
def RGB2HEX(color):

    """convert RGB value to HEX format
    
    Inputs: 
    
        color: color in rgb format
        
    Returns:
    
        color in hex format
        
    """   
    return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2]))



# get the color pallate
def get_cmap(n, name = 'hsv'):

    """get n kinds of colors from a color palette 
    
    Inputs: 
    
        n: number of colors
        
        name: the color palette choosed
        
    Returns:
    
        plt.cm.get_cmap(name, n): Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
        RGB color; the keyword argument name must be a standard mpl colormap name. 
        
    """   
    return plt.cm.get_cmap(name, n)
    


def barcode_detect(img_ori):
    
    """Read barcode in the image and decode barcode info
    
    Inputs: 
    
        img_ori: image contains the barcode region
        
    Returns:
    
        tag_info: decoded barcode information  
        
    """
    
    # get the dimension of the image
    height, width = img_ori.shape[:2]
    
    # decode the barcode info 
    barcode_info = decode((img_ori.tobytes(), width, height))
    
    # if barcode info was not empty
    if len(barcode_info) > 0:
        
        # get the decoded barcode info data value as string
        barcode_str = str(barcode_info[0].data)
        
        #print('Decoded data:', barcode_str)
        #print(decoded_object.rect.top, decoded_object.rect.left)
        #print(decoded_object.rect.width, decoded_object.rect.height)
 
        # accquire the barcode info and remove extra characters
        tag_info = re.findall(r"'(.*?)'", barcode_str, re.DOTALL)
        tag_info = " ".join(str(x) for x in tag_info)
        tag_info = tag_info.replace("'", "")
        
        print("Tag info: {}\n".format(tag_info))
    
    else:
        # print warning if barcode info was empty
        print("barcode information was not readable!\n")
        tag_info = 'Unreadable'
        
    return tag_info
    
    


def marker_detect(img_ori, template, method, selection_threshold):
    
    """Detect marker in the image
    
    Inputs: 
    
        img_ori: image contains the marker region
        
        template: preload marker template image
        
        method: method used to compute template matching
        
        selection_threshold: thresh value for accept the template matching result

    Returns:
    
        marker_img: matching region image with marker object  
        
        thresh: mask image of the marker region
        
        coins_width_contour: computed width result based on contour of the object 
        
        coins_width_circle: computed width result based on min circle of the object 
        
    """   
    
    # load the image, clone it for output
    img_rgb = img_ori.copy()
      
    # convert it to grayscale 
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY) 
      
    # store width and height of template in w and h 
    w, h = template.shape[::-1] 
      
    # Perform match operations. 
    #res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
    
    # Perform template matching operations. 
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF)
    
    #res = cv2.matchTemplate(img_gray, template, cv2.TM_SQDIFF)
    
    
    # Specify a threshold for template detection as selection_threshold
    
    # Store the coordinates of matched area in a numpy array 
    loc = np.where( res >= selection_threshold)   
    
    if len(loc):
        
        # unwarp the template mathcing result
        (y,x) = np.unravel_index(res.argmax(), res.shape)
        
        # get the template matching region coordinates
        (min_val, max_val, min_loc, max_loc) = cv2.minMaxLoc(res)

        (startX, startY) = max_loc
        endX = startX + template.shape[1]
        endY = startY + template.shape[0]
        
        # get the sub image with matching region
        marker_img = img_ori[startY:endY, startX:endX]
        marker_overlay = marker_img

        # load the marker image, convert it to grayscale
        marker_img_gray = cv2.cvtColor(marker_img, cv2.COLOR_BGR2GRAY) 

        # load the image and perform pyramid mean shift filtering to aid the thresholding step
        shifted = cv2.pyrMeanShiftFiltering(marker_img, 21, 51)
        
        # convert the mean shift image to grayscale, then apply Otsu's thresholding
        gray = cv2.cvtColor(shifted, cv2.COLOR_BGR2GRAY)
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        
        
        # detect contours in the mask and grab the largest one
        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        cnts = imutils.grab_contours(cnts)
        
        largest_cnt = max(cnts, key=cv2.contourArea)
        
        #print("[INFO] {} unique contours found in marker_img\n".format(len(cnts)))
        
        # compute the radius of the detected coin
        # calculate the center of the contour
        M = cv2.moments(largest_cnt )
        x = int(M["m10"] / M["m00"])
        y = int(M["m01"] / M["m00"])

        # calculate the radius of the contour from area (I suppose it's a circle)
        area = cv2.contourArea(largest_cnt)
        radius = np.sqrt(area/math.pi)
        coins_width_contour = 2* radius
    
        # draw a circle enclosing the object
        ((x, y), radius) = cv2.minEnclosingCircle(largest_cnt) 
        coins_width_circle = 2* radius
    
    else:
        
        print("no matching template was found\n")

    return  marker_img, thresh, coins_width_contour, coins_width_circle



def adjust_gamma(image, gamma):

    """Adjust the gamma value to increase the brightness of image
    
    Inputs: 
    
        image: image 
        
        gamma: gamma value used to adjust

    Returns:
    
        cv2.LUT(image, table): adjusted image with gamma correction
        
        
    """
    
    # build a lookup table mapping the pixel values [0, 255] to
    # their adjusted gamma values
    invGamma = 1.0 / gamma
    table = np.array([((i / 255.0) ** invGamma) * 255
        for i in np.arange(0, 256)]).astype("uint8")
 
    # apply gamma correction using the lookup table
    return cv2.LUT(image, table)



def closest_center(pt, pt_list):
    
    """compute index of closest point between current point and a list of points 
    
    Inputs: 
    
        pt: coordinate of current point
        
        pt_list: coordinates of a list of points

    Returns:
    
        min_dist_index: index of closest point
        
    """
    min_dist_index = np.argmin(np.sum((np.array(pt_list) - np.array(pt))**2, axis=1))
    
    return min_dist_index
    


def circle_detection(image):

    """Detecting Circles in Images using OpenCV and Hough Circles
    
    Inputs: 
    
        image: image loaded 

    Returns:
    
        circles: detcted circles
        
        circle_detection_img: circle overlayed with image
        
        diameter_circle: diameter of detected circle
        
    """
    
    # create background image for drawing the detected circle
    output = image.copy()
    
    # obtain image dimension
    img_height, img_width, n_channels = image.shape
    
    #backup input image
    circle_detection_img = image.copy()
    
    # change image from RGB to Gray scale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # apply blur filter 
    blurred = cv2.medianBlur(gray, 25)
    
    # setup parameters for circle detection
    
    # This parameter is the inverse ratio of the accumulator resolution to the image resolution 
    #(see Yuen et al. for more details). Essentially, the larger the dp gets, the smaller the accumulator array gets.
    dp = 1.5
    
    #Minimum distance between the center (x, y) coordinates of detected circles. 
    #If the minDist is too small, multiple circles in the same neighborhood as the original may be (falsely) detected. 
    #If the minDist is too large, then some circles may not be detected at all.
    minDist = 100
    
    #Gradient value used to handle edge detection in the Yuen et al. method.
    #param1 = 30
    
    #accumulator threshold value for the cv2.HOUGH_GRADIENT method. 
    #The smaller the threshold is, the more circles will be detected (including false circles). 
    #The larger the threshold is, the more circles will potentially be returned. 
    #param2 = 30  
    
    #Minimum/Maximum size of the radius (in pixels).
    #minRadius = 80
    #maxRadius = 120 
    
    # detect circles in the image
    #circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, 1.2, minDist, param1=param1, param2=param2, minRadius=minRadius, maxRadius=maxRadius)
    
    # detect circles in the image
    circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, dp, minDist)
    
    # initialize diameter of detected circle
    diameter_circle = 0
    
    
    circle_center_coord = []
    circle_center_radius = []
    idx_closest = 0
    
    if circles is not None: 
        # convert the (x, y) coordinates and radius of the circles to integers
        circles = np.round(circles[0, :]).astype("int")
        
        # loop over the (x, y) coordinates and radius of the circles
        for (x, y, r) in circles:
            
            coord = (x, y)
            
            circle_center_coord.append(coord)
            circle_center_radius.append(r)
        
        
        # choose the left bottom circle if more than one circles are detected 
        if len(circles) > 1:
            
            #finding closest point among the center list of the circles to the right-bottom of the image
            idx_closest = closest_center((0 + img_width, 0 + img_height), circle_center_coord)
        
        else:
            
            # ensure at least some circles were found
            if circles is not None and len(circles) > 0:
                idx_closest = 0
    
        print("idx_closest = {}\n".format(idx_closest))
        
        # draw the circle in the output image, then draw a center
        circle_detection_img = cv2.circle(output, circle_center_coord[idx_closest], circle_center_radius[idx_closest], (0, 255, 0), 4)
        circle_detection_img = cv2.circle(output, circle_center_coord[idx_closest], 5, (0, 128, 255), -1)

        # compute the diameter of coin
        diameter_circle = circle_center_radius[idx_closest]*2
    
    
    return circles, circle_detection_img, diameter_circle
        



# compute rect region based on left top corner coordinates and dimension of the region
def region_extracted(orig, x, y, w, h):
    
    """compute rect region based on left top corner coordinates and dimension of the region
    
    Inputs: 
    
        orig: image
        
        x, y: left top corner coordinates 
        
        w, h: dimension of the region

    Returns:
    
        roi: region of interest
        
    """   
    roi = orig[y:y+h, x:x+w]
    
    return roi





def get_marker_region(orig, mask_external):
    
    """compute masked image for coin and tag detection.
    
    Inputs: 
    
        orig: image data
        
        mask_external: detected mask for foreground objects

    Returns:
    
        roi_image: masked image by filling the foreground objects by black color
        
    """
    # create an size 10 kernel
    kernel = np.ones((25,25), np.uint8)
    
    # image dilation
    dilation = cv2.dilate(mask_external.copy(), kernel, iterations = 1)
    
    # image closing
    mask_external_dilate = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
    
    
    # Inverting the mask by performing bitwise-not operation
    mask_external_invert = cv2.bitwise_not(mask_external_dilate)
    
    roi_image = cv2.bitwise_and(orig, orig, mask = mask_external_invert)
    

    return roi_image





def isbright(orig):
    
    """compute the brightness of the input image, Convert it to LAB color space to access the luminous channel which is independent of colors.
    
    Inputs: 
    
        orig: image data after loading

    Returns:
    
        np.mean(L): brightness value of the image
        
    """
    
    # Set up threshold value for luminous channel, can be adjusted and generalized 
    thresh = 1.5
    
    # Make backup image
    image = orig.copy()
    
    
    # Convert color space to LAB format and extract L channel
    L, A, B = cv2.split(cv2.cvtColor(image, cv2.COLOR_BGR2LAB))
    
    # Normalize L channel by dividing all pixel values with maximum pixel value
    L = L/np.max(L)
    
    text_bool = "bright" if np.mean(L) < thresh else "dark"
    
    #return image_file_name, np.mean(L), text_bool
    
    #print("np.mean(L) < thresh = {}".format(np.mean(L)))
    
    #return np.mean(L) < thresh
    
    return np.mean(L)



def watershed_seg(orig, min_distance_value):

    """segment individual connected / overlaped object based on wastershed segmentation method
    
    Inputs: 
    
        orig: masked image contains only target objects
        
        min_distance_value: min distance between each peaks in the distance map

    Returns:
    
        labels: matrix, Each pixel value as a unique label value. Pixels that have the same label value belong to the same object.
        
        label_overlay: overlay original image with all labels
        
        labeled_img: label image in hue map format
        
        count_kernel: count of the segmented kernels 
        
    """
   
    image = orig.copy()
     
    # convert the mean shift image to grayscale, then apply Otsu's thresholding
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    # compute the exact Euclidean distance from every binary pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    
    
    localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
    
    markers = ndimage.label(localMax, structure=np.ones((3, 3)))[0]
    
    labels = watershed(-D, markers, mask=thresh)
    
    
    '''
    max_coords = peak_local_max(D, min_distance = min_distance_value, labels = thresh)
    
    local_maxima = np.zeros_like(image, dtype=bool)
    
    local_maxima[tuple(max_coords.T)] = True
    
    # perform a connected component analysis on the local peaks, using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(local_maxima)[0]  
    
   
    #print("markers")
    #print(type(markers))
    
    labels = watershed(-D, markers, mask = img_as_float(thresh))
    
    from skimage.color import label2rgb
    labeled_img = label2rgb(labeled_coins, image=coins)
    '''
    print("[INFO] {} unique labels found\n".format(len(np.unique(labels)) - 1))
   
    #Map component labels to hue val
    label_hue = np.uint8(128*labels/np.max(labels))
    blank_ch = 255*np.ones_like(label_hue)
    labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

    # cvt to BGR for display
    labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

    # set background label to black
    labeled_img[label_hue==0] = 0
   
    #define result path for labeled images
    #result_img_path = save_path_label + str(filename[0:-4]) + '_label.jpg'

    # save results
    #cv2.imwrite(result_img_path,labeled_img)
    area_rec = []
    kernel_size = 0
    count_kernel = 0
    # loop over the unique labels returned by the Watershed algorithm
    for label in np.unique(labels):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue
     
        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros(gray.shape, dtype="uint8")
        mask[labels == label] = 255
        
  
        # detect contours in the mask and grab the largest one
        cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)
        c = max(cnts, key=cv2.contourArea)
     
        # draw a circle enclosing the object
        ((x, y), r) = cv2.minEnclosingCircle(c)
        if r > 0:
            label_overlay = cv2.circle(image, (int(x), int(y)), 2, (255, 0, 0), 5)
            label_overlay = cv2.drawContours(image, [c], -1, (0, 255, 0), 2)
            #cv2.putText(image, "#{}".format(label), (int(x) - 10, int(y)),cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
            count_kernel+= 1
            area_rec.append(cv2.contourArea(c))
            

    

   
    kernel_size = sum(area_rec)/len(area_rec)
    
    max_kernel_size = max(area_rec)
    
    min_kernel_size = min(area_rec)
    
    print("[INFO] segments number = {}, size = {}\n".format(count_kernel, kernel_size))
    
    return labels, label_overlay, labeled_img, count_kernel, kernel_size, max_kernel_size, min_kernel_size



   
'''
def adaptive_threshold(masked_image, GaussianBlur_ksize, blockSize, weighted_mean):
    
    """compute thresh image using adaptive threshold Method
    
    Inputs: 
    
        maksed_img: masked image contains only target objects
        
        GaussianBlur_ksize: Gaussian Kernel Size 
        
        blockSize: size of the pixelneighborhood used to calculate the threshold value
        
        weighted_mean: the constant used in the both methods (subtracted from the mean or weighted mean).

    Returns:
        
        thresh_adaptive_threshold: thresh image using adaptive thrshold Method
        
        maksed_img_adaptive_threshold: masked image using thresh_adaptive_threshold

    """
    ori = masked_image.copy()
    
    if len(ori.shape)> 2:
        # convert the image to grayscale and blur it slightly
        gray = cv2.cvtColor(masked_image, cv2.COLOR_BGR2GRAY)
    else:
        gray = ori
    
    # blurring it . Applying Gaussian blurring with a GaussianBlur_ksize×GaussianBlur_ksize kernel 
    # helps remove some of the high frequency edges in the image that we are not concerned with and allow us to obtain a more “clean” segmentation.
    blurred = cv2.GaussianBlur(gray, (GaussianBlur_ksize, GaussianBlur_ksize), 0)

    # adaptive method to be used. 'ADAPTIVE_THRESH_MEAN_C' or 'ADAPTIVE_THRESH_GAUSSIAN_C'
    thresh_adaptive_threshold = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, blockSize, weighted_mean)

    # apply individual object mask
    maksed_img_adaptive_threshold = cv2.bitwise_and(ori, ori.copy(), mask = ~thresh_adaptive_threshold)

    return thresh_adaptive_threshold, maksed_img_adaptive_threshold

'''
'''

def kernel_traits_computation(masked_img, labels):

    """compute kernel traits based on input image and its segmentation labels
    
    Inputs: 
    
        masked_img: masked image contains only target objects
        
        labels: watershed_seg return matrix, Each pixel value as a unique label value. Pixels that have the same label value belong to the same object.
                

    Returns:
        
        label_trait: overlay image with all traits visualization
        
        kernel_index_rec: index of each kernel.
        
        contours_rec: list of contours of each kernel.
        
        area_rec: list of area of each kernel.
        
        major_axis_rec: list of major_axis of each kernel. (Bounding box)
        
        minor_axis_rec: list of minor_axis of each kernel. (Bounding box)
        
        
    """
    # initialize parameters
    orig = masked_img.copy()
    
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    
    kernel_index_rec = []
    contours_rec = []
    area_rec = []

    major_axis_rec = []
    minor_axis_rec = []
    
    count = 0


    # loop over the unique labels returned by the Watershed algorithm
    for index, label in enumerate(np.unique(labels), start = 1):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue
     
        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros(gray.shape, dtype = "uint8")
        mask[labels == label] = 255
        
        # apply individual object mask
        masked = cv2.bitwise_and(orig, orig, mask = mask)
        
        #individual kernel segmentation 
        
        # detect contours in the mask and grab the largest one
        #cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
        contours, hierarchy = cv2.findContours(mask.copy(),cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        c = max(contours, key = cv2.contourArea)

        if len(c) >= 10 :

            contours_rec.append(c)
            area_rec.append(cv2.contourArea(c))

        #else:
            # optional to "delete" the small contours
            #print("small contours")
    
 
    # sort the contours based on area size order
    contours_rec_sorted = [x for _, x in sorted(zip(area_rec, contours_rec), key=lambda pair: pair[0])]
    
    #cmap = get_cmap(len(contours_rec_sorted)) 
    
    cmap = get_cmap(len(contours_rec_sorted)+1)
    
    
    #tracking_backgd = np.zeros(gray.shape, dtype = "uint8")
    #backgd.fill(128)
    
    label_trait = orig.copy()
    
    #track_trait = orig.copy()
    #clean area record list
    area_rec = []
    #individual kernel traits sorting based on area order 
    ################################################################################
    for i in range(len(contours_rec_sorted)):
        
        c = contours_rec_sorted[i]
        
        #assign unique color value in opencv format
        color_rgb = tuple(reversed(cmap(i)[:len(cmap(i))-1]))
        
        color_rgb = tuple([255*x for x in color_rgb])
        
        
        # get coordinates of bounding box
        
        (x,y,w,h) = cv2.boundingRect(c)
    
        
        
        # draw a circle enclosing the object
        ((x, y), r) = cv2.minEnclosingCircle(c)
        #label_trait = cv2.circle(orig, (int(x), int(y)), 3, (0, 255, 0), 2)
        
        #draw filled contour
        #label_trait = cv2.drawContours(orig, [c], -1, color_rgb, -1)
        if cv2.contourArea(c) < 6523:
            
            label_trait = cv2.drawContours(orig, [c], -1, color_rgb, 2)
        
        #label_trait = cv2.putText(orig, "#{}".format(i+1), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color_rgb, 1)
        
        #label_trait = cv2.putText(backgd, "#{}".format(i+1), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color_rgb, 1)
        
        #draw mini bounding box
        #label_trait = cv2.drawContours(orig, [box], -1, (0, 255, 0), 2)
        
        #######################################individual kernel curvature computation
        
        #record all traits 
        kernel_index_rec.append(i)
        area_rec.append(cv2.contourArea(c))
        #curv_rec.append(curvature)
        

        major_axis_rec.append(w)
        minor_axis_rec.append(h)
        
    ################################################################################
    
    
    #print('unique labels={0}, len(contours_rec)={1}, len(kernel_index_rec)={2}'.format(np.unique(labels),len(contours_rec),len(kernel_index_rec)))
        
    n_contours = len(contours_rec_sorted)
    

    return label_trait, kernel_index_rec, contours_rec, area_rec, major_axis_rec, minor_axis_rec
    
'''

def valid_kernel_mask(orig_mask, cnt_width, cnt_height, cnt_x, cnt_y, valid_kernel_ratio_list):
    
    """compute mask image for valid kernel area defined by the ratio of top/bottom to the ear length
    
    Inputs: 
    
        orig_mask: ear object segmentation mask
        
        cnt_width, cnt_height: dimension of the ear objects
        
        cnt_x, cnt_y: coordinates of ear objects
        
        valid_kernel_ratio_top, valid_kernel_ratio_bottom: user defined ratio of top/bottom to the ear length

    Returns:
    
        v_mask: mask image for valid kernel area
        
    """
    
    # extract ratio values 
    valid_kernel_ratio_left = valid_kernel_ratio_list[0]
    valid_kernel_ratio_right = valid_kernel_ratio_list[1]
    valid_kernel_ratio_top = valid_kernel_ratio_list[2]
    valid_kernel_ratio_bottom = valid_kernel_ratio_list[3]
    
    # a mask with the same size as inout image, pixels with a value of 0 (background) are
    # ignored in the original image while mask pixels with a value of
    # 255 (foreground) are allowed to be kept
    img_height, img_width = orig_mask.shape
    
    #print(img_height, img_width)
    
    # initialize empty image as mask
    v_mask = np.zeros(orig_mask.shape, dtype = "uint8")
    
    # compute the coordinates to get the masking area
    #x_l = 0
    #x_r = int(avg_x + img_width)
    #y_t = int(avg_y + max_height*valid_kernel_ratio_top)
    #y_b = int(avg_y + max_height*(1-valid_kernel_ratio_bottom))
    '''
    # compute the coordinates to get the masking area for two ears in one image
    for i in range(len(cnt_height)):
    
        # compute the coordinates to get the masking area
        x_l = int(cnt_x[i] + (cnt_width[i]-10)*valid_kernel_ratio_left)
        x_r = int(cnt_x[i] + (cnt_width[i]-10)*(1 - valid_kernel_ratio_right))
        y_t = int(cnt_y[i] + (cnt_height[i]-10)*valid_kernel_ratio_top)
        y_b = int(cnt_y[i] + (cnt_height[i]-10)*(1-valid_kernel_ratio_bottom))
    
        # assign area of valid kernel 
        v_mask[y_t : y_b, x_l : x_r] = 255
    '''
    
    # compute the coordinates to get the masking area
    x_l = int(cnt_x + (cnt_width-10)*valid_kernel_ratio_left)
    x_r = int(cnt_x + (cnt_width-10)*(1 - valid_kernel_ratio_right))
    y_t = int(cnt_y + (cnt_height-10)*valid_kernel_ratio_top)
    y_b = int(cnt_y + (cnt_height-10)*(1-valid_kernel_ratio_bottom))

    # assign area of valid kernel 
    v_mask[y_t : y_b, x_l : x_r] = 255
    
    return v_mask
    




def get_contours(image_thresh):
    
    """find contours in binary image mask and sort them in left-to-right order
    
    Inputs: 
    
        image_thresh: image mask

    Returns:
    
        cnts_sorted:  sorted contours
        
    """
    # find contours in the thresholded image
    cnts = cv2.findContours(image_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    cnts = imutils.grab_contours(cnts)

    # sort the contour based on area size from largest to smallest, and get the first two max contours
    cnts_sorted = sorted(cnts, key = cv2.contourArea, reverse = True)[0:n_ear]

    # sort the contours from left to right
    cnts_sorted = sort_contours(cnts_sorted, method = "left-to-right")

   
    print("Sorting {} objects in left-to-right order\n".format(len(cnts_sorted)))
    
    return cnts_sorted





def mask_from_contour(orig, c):
    
    """compute binary mask image  from contours
    
    Inputs: 
    
        img_thresh: image mask
        
        

    Returns:
    
        cnts_sorted:  sorted contours
        
    """
    
    # initialize empty mask image
    img_thresh = np.zeros(orig.shape, np.uint8)
        
    mask_individual_BRG = cv2.drawContours(img_thresh, [c], -1, (255, 255, 255), -1)

    #v_mask = valid_kernel_mask(mask_internal.copy(), w, h, x, y, valid_kernel_ratio_list)

    mask_individual_gray = cv2.cvtColor(mask_individual_BRG, cv2.COLOR_BGR2GRAY)

    mask_individual = cv2.threshold(mask_individual_gray, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

    return mask_individual






def kernel_traits_computation(masked_kernel, ):
    
    """analyze kernel traits in maize ear  
    
    Inputs: 
    
        masked_kernel: maksed individual ear image
        
        

    Returns:
    
        labels_kernel: kernel labels in segmentation results
        
        label_overlay: visualization of segmentation labels
        
        labeled_img: image of visualization of segmentation labels
        
        count_kernel: number of kernels
        
        kernel_size: average size of kernels
        
    """

    ###################################################################################
    # set the parameters for adoptive threshholding method
    GaussianBlur_ksize = 5

    blockSize = 41

    weighted_mean = 10

    # adoptive threshholding method to the masked image from mutilple_objects_seg
    (thresh_adaptive_threshold, maksed_img_adaptive_threshold) = adaptive_threshold(masked_kernel, GaussianBlur_ksize, blockSize, weighted_mean)

    # save result
    #result_file = (save_path + base_name + '_thresh_adaptive_threshold' + file_extension)
    #cv2.imwrite(result_file, thresh_adaptive_threshold)

    # save result
    #result_file = (save_path + base_name + '_maksed_img_adaptive_threshold' + file_extension)
    #cv2.imwrite(result_file, maksed_img_adaptive_threshold)

    #using wahtershed method to segement the kernels
    (labels, label_overlay, labeled_img, count_kernel, kernel_size, max_kernel_size, min_kernel_size) = watershed_seg(maksed_img_adaptive_threshold, min_distance_value)
    
    
    return labels, label_overlay, labeled_img, count_kernel, kernel_size, max_kernel_size, min_kernel_size
    



def extract_traits(image_file):

    """compute all the traits based on input image
    
    Inputs: 
    
        image file: full path and file name

    Returns:
        image_file_name: file name
        
        tag_info: Barcode information
        
        tassel_area: area occupied by the tassel in the image
        
        tassel_area_ratio: The ratio between tassel area and its convex hull area
        
        cnt_width, cnt_height: width and height of the tassel
        
        n_branch: number of branches in the tassel
        
        avg_branch_length: average length of the branches
        
        branch_length: list of all branch length
    """
    
    # extarct path and name of the image file
    abs_path = os.path.abspath(image_file)
    
    # extract the base name 
    filename, file_extension = os.path.splitext(abs_path)
    base_name = os.path.splitext(os.path.basename(filename))[0]

    # get the file size
    file_size = os.path.getsize(image_file)/MBFACTOR
    
    # get the image file name
    image_file_name = Path(image_file).name


    print("Extracting traits for image : {0}\n".format(str(image_file_name)))
     
    # create result folder
    if (args['result']):
        save_path = args['result']
    else:
        mkpath = os.path.dirname(abs_path) +'/' + base_name
        mkdir(mkpath)
        save_path = mkpath + '/'
        
    print ("results_folder:" + save_path +'\n')

    


    
    if (file_size > 5.0):
        print("File size: {0} MB\n".format(str(int(file_size))))
    else:
        print("Plant object segmentation using automatic color clustering method...\n")
    
    
    # load the input image 
    image = cv2.imread(image_file)

    #make backup image
    orig = image.copy()
    
    # compute image brightness value to record illumination condition of images
    img_brightness = isbright(orig)
    #print ("image brightness is {}\n".format(img_brightness)) 
    
    # compute image blurriness value to record images out of focus
    #blurry_value = 0
    #blurry_value = detect_blur(orig)
    
    #print("Image blurry value: {0}\n".format(blurry_value))
    
    # get the dimension of the image
    img_height, img_width, img_channels = orig.shape
    
    #source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)
    
    ##########################################################################################################
    # segment mutiple objects in image uto accquire external contours
    (mask_external_ai, img_foreground) = mutilple_objects_detection(orig)
    
    (mask_external_cluster) = mutilple_objects_seg(orig, channel = 'L', size_kernel = 5)
    
   
    #color clustering based object segmentation to accquire another external contours
    #mask_external_cluster = color_cluster_seg(image.copy(), args['color_space'], args['channels'], args['num_clusters'])
    
    mask_external_combined = mask_external_ai & mask_external_cluster
    
    #mask_external_combined = mask_external_ai 
    
    mask_external = cv2.threshold(mask_external_combined, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
        
    # apply individual object mask
    img_foreground = cv2.bitwise_and(image.copy(), image.copy(), mask = mask_external_combined)
    
    
    # save result
    result_file = (save_path + base_name + '_mask_external' + file_extension)
    cv2.imwrite(result_file, mask_external)

    # save result
    result_file = (save_path + base_name + '_foreground' + file_extension)
    cv2.imwrite(result_file, img_foreground)
    
    
    
    
    ###########################################################################################################
    # segment mutiple objects in image using thresh method to accquire internal contours
    #(left_img, right_img, mask_seg, img_overlay, cnt_area_internal) = mutilple_objects_seg(orig, channel = 'B')
    
    (mask_internal) = mutilple_objects_seg(orig, channel = 'B', size_kernel = 10)
    
    #mask_internal = mask_seg
    
    # apply individual object mask
    #masked_image = cv2.bitwise_and(image.copy(), image.copy(), mask = mask_seg)
    
    masked_image = img_foreground
    
    # save result
    result_file = (save_path + base_name + '_mask_internal' + file_extension)
    cv2.imwrite(result_file, mask_internal)
    


   ###############################################################################################
    #combine external contours and internal contours to compute object mask
    #combined_mask = mask_seg | thresh | mask_external
    
    #combined_mask = mask_internal | mask_external

    
    #thresh_combined_mask = cv2.threshold(mask_external, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    # initialize all the traits output 
    area = kernel_area_ratio = max_width = max_height = avg_curv = 0
    
    n_kernels_valid = kernel_size = n_kernels_all = kernel_area = 0
    
    
    ################################
    # compute triats for every kernel in current image
    ear_area_rec = []
    ear_area_ratio_rec = []
    
    ear_width_rec = []
    ear_height_rec = []
    
    n_kernels_all_rec = []
    n_kernels_valid_rec = []
    
    kernel_size_rec = []
    kernel_valid_size_rec = []
    
    ear_index_rec = []
    
    
    # get contours from both exteanl and internal masks
    cnts_external = get_contours(mask_external)
    cnts_internal = get_contours(mask_internal)

    
    # set the parameters for wateshed segmentation method
    min_distance_value = args['min_dist']
    
    # initialize background image to draw the contours
    img_overlay_bk = orig.copy()
    
    # loop over the selected contours
    for idx, (c_external, c_internal) in enumerate(zip(cnts_external, cnts_internal)):
        
        
        ear_index_rec.append(idx+1)
        ##########################################################################################
        #get the bounding rect
        (x, y, w, h) = cv2.boundingRect(c_external)
        
        area_c = max(cv2.contourArea(c_external), cv2.contourArea(c_internal))
        
        area_ratio = min(cv2.contourArea(c_external), cv2.contourArea(c_internal))/area_c
        
        ear_area_rec.append(area_c)
        ear_width_rec.append(w)
        ear_height_rec.append(h)
        
        ear_area_ratio_rec.append(area_ratio)
        
        print("Contour shape info: width = {1:.2f}, height = {2:.2f}, area = {3:.2f}, area_ratio = {3:.2f}\n".format(w, h, area_c, area_ratio))
        ############################################################################################

        mask_individual = mask_from_contour(orig.copy(), c_internal)
        
        # apply individual object mask
        masked_kernel = cv2.bitwise_and(img_foreground.copy(), img_foreground.copy(), mask = mask_individual)
        
        # save result
        #result_file = (save_path + base_name + str("_{}".format(idx)) +  '_masked_kernel' + file_extension)
        #cv2.imwrite(result_file, masked_kernel)
        
        
        (labels_kernel, label_overlay, labeled_img, count_kernel, kernel_size, max_kernel_size, min_kernel_size) = kernel_traits_computation(masked_kernel)
        

        n_kernels_all_rec.append(count_kernel)
        kernel_size_rec.append(kernel_size)
        

        # save result
        #result_file = (save_path + base_name + str("_{}".format(idx)) + '_label_overlay' + file_extension)
        #cv2.imwrite(result_file, label_overlay)
        

        
        ###############################################################################################
        v_mask = valid_kernel_mask(mask_individual.copy(), w, h, x, y, valid_kernel_ratio_list)
        
        # apply individual object mask
        masked_kernel_valid = cv2.bitwise_and(img_foreground.copy(), img_foreground.copy(), mask = v_mask)
        
        # save result
        #result_file = (save_path + base_name + str("_{}".format(idx)) +  '_masked_valid_kernel' + file_extension)
        #cv2.imwrite(result_file, masked_kernel_valid)
        
        
        (labels_kernel_valid, label_overlay_valid, labeled_img_valid, count_kernel_valid, kernel_size_valid, max_kernel_size_valid, min_kernel_size_valid) = kernel_traits_computation(masked_kernel_valid)
        
        n_kernels_valid_rec.append(count_kernel_valid)
        kernel_valid_size_rec.append(kernel_size_valid)
        
        # save result
        #result_file = (save_path + base_name + str("_{}".format(idx)) + '_label_overlay_valid' + file_extension)
        #cv2.imwrite(result_file, label_overlay_valid)
        

        ############################################################################################
        # visualize results
        
        # compute the center of the contour
        M = cv2.moments(c_external)
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        
        # draw the center of the shape on the image
        trait_img = cv2.circle(img_overlay_bk, (cX, cY), 7, (255, 255, 255), -1)
        #trait_img = cv2.putText(img_overlay_bk, "center", (cX - 20, cY - 20),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
            
        
        # compute the four coordinates to get the center of bounding box
        tl = (x, y+h*0.5)
        tr = (x+w, y+h*0.5)
        br = (x+w*0.5, y)
        bl = (x+w*0.5, y+h)
        
        # compute the midpoint between bottom-left and bottom-right coordinates
        (tltrX, tltrY) = midpoint(tl, tr)
        (blbrX, blbrY) = midpoint(bl, br)
        
        # draw the midpoints on the image
        trait_img = cv2.circle(img_overlay_bk, (int(tltrX), int(tltrY)), 15, (255, 0, 0), -1)
        trait_img = cv2.circle(img_overlay_bk, (int(blbrX), int(blbrY)), 15, (255, 0, 0), -1)

        # draw lines between the midpoints
        trait_img = cv2.line(img_overlay_bk, (int(x), int(y+h*0.5)), (int(x+w), int(y+h*0.5)), (255, 0, 255), 6)
        trait_img = cv2.line(img_overlay_bk, (int(x+w*0.5), int(y)), (int(x+w*0.5), int(y+h)), (255, 0, 255), 6)
        
        # compute the convex hull of the contour
        hull = cv2.convexHull(c_external)
        
        # draw convexhull in red color
        trait_img = cv2.drawContours(img_overlay_bk, [hull], -1, (0, 0, 255), 2)
        
        
        # draw a green rectangle to visualize the bounding rect
        trait_img = cv2.rectangle(img_overlay_bk, (x, y), (x+w, y+h), (255, 255, 0), 4)
            
        # draw the contour and center of the shape on the image
        trait_img = cv2.drawContours(img_overlay_bk, [c_external], -1, (0, 255, 0), 2)
        
        # draw the contour and center of the shape on the image
        trait_img = cv2.drawContours(img_overlay_bk, [c_internal], -1, (0, 0, 255), 2)

        #center_result = cv2.circle(img_thresh, (cX, cY), 14, (0, 0, 255), -1)
        trait_img = cv2.putText(img_overlay_bk, "{}".format(idx +1), (cX - 20, cY - 20), cv2.FONT_HERSHEY_SIMPLEX, 2.5, (255, 0, 0), 5)
    
    
    # save result
    result_file = (save_path + base_name + '_overlay' + file_extension)
    cv2.imwrite(result_file, trait_img)
    
    
    
    ###################################################################################################
    # detect coin and barcode uisng template matching and circle detection method
    
    
    roi_image = get_marker_region(orig, mask_external)
    
    # save result
    #result_file = (save_path + base_name + '_coin_region' + file_extension)
    #cv2.imwrite(result_file, roi_image)
    
    # apply gamma correction for image region with coin
    gamma = 1.5
    gamma = gamma if gamma > 0 else 0.1
    enhanced_region = adjust_gamma(roi_image.copy(), gamma = gamma)
    
    (circles, circle_detection_img, diameter_circle) = circle_detection(enhanced_region) 
    
    if diameter_circle > 0 :
        pixel_cm_ratio = diameter_circle/coin_size
    else:
        pixel_cm_ratio = 1
    
    # save result
    result_file = (save_path + base_name + '_coin_circle' + file_extension)
    cv2.imwrite(result_file, circle_detection_img)
    
    print("The width of coin in the marker image is {:.0f} × {:.0f} pixels\n".format(diameter_circle, diameter_circle))
    
    
    
    # barcode detection
    # method for template matching 
    methods = ['cv.TM_CCOEFF', 'cv.TM_CCOEFF_NORMED', 'cv.TM_CCORR',
        'cv.TM_CCORR_NORMED', 'cv.TM_SQDIFF', 'cv.TM_SQDIFF_NORMED']
    
    # apply gamma correction for image region with coin
    gamma = 1.5
    gamma = gamma if gamma > 0 else 0.1
    enhanced_region = adjust_gamma(roi_image.copy(), gamma=gamma)
    
    # detect the barcode object based on template image
    (marker_barcode_img, thresh_barcode, barcode_width_contour, barcode_width_circle) = marker_detect(enhanced_region, tp_barcode, methods[0], 0.8)
    
    # parse barcode image using pylibdmtx lib
    tag_info = barcode_detect(marker_barcode_img)

    if tag_info != 'Unreadable':
        # save result
        result_file = (save_path + base_name + '_barcode' + file_extension)
        cv2.imwrite(result_file, marker_barcode_img)
    
    #return image_file_name, tag_info, kernel_size, n_kernels_valid, n_kernels_all, kernel_area, kernel_area_ratio, max_width, max_height, diameter_circle, coin_size, pixel_cm_ratio, img_brightness, blurry_value
    
    return image_file_name, tag_info, ear_index_rec, ear_area_rec, ear_area_ratio_rec, ear_width_rec, ear_height_rec,\
            n_kernels_all_rec, kernel_size_rec, n_kernels_valid_rec, kernel_valid_size_rec, \
            diameter_circle, coin_size, pixel_cm_ratio




if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", required = True,    help = "Image filetype")
    ap.add_argument('-mk', '--marker', required = False,  default ='/marker_template/coin.png',  help = "Marker file name")
    ap.add_argument('-bc', '--barcode', required = False,  default ='/marker_template/barcode.png',  help = "Barcode file name")
    ap.add_argument("-r", "--result", required = False,    help="result path")
    ap.add_argument('-s', '--color-space', type = str, required = False, default ='Lab', help='Color space to use: BGR , HSV, Lab(default), YCrCb (YCC)')
    ap.add_argument('-c', '--channels', type = str, required = False, default='0', help='Channel indices to use for clustering, where 0 is the first channel,' 
                                                                       + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" ' 
                                                                       + 'selects channels B and R. (default "all")')
    ap.add_argument('-n', '--num-clusters', type = int, required = False, default = 2,  help = 'Number of clusters for K-means clustering (default 2, min 2).')
    ap.add_argument('-ne', '--num_ears', type = int, required = False, default = 2,  help = 'Number of ears in image (default 2).')
    ap.add_argument('-min', '--min_size', type = int, required = False, default = 250000,  help = 'min size of object to be segmented.')
    ap.add_argument('-md', '--min_dist', type = int, required = False, default = 4,  help = 'distance threshold for watershed segmentation.')
    ap.add_argument('-cs', '--coin_size', type = int, required = False, default = 2.7,  help = 'coin diameter in cm')
    ap.add_argument('-vkrl', '--valid_kernel_ratio_left', type = float, required = False, default = 0.10,  help = 'valid kernel ratio copmpared with ear width from left')
    ap.add_argument('-vkrr', '--valid_kernel_ratio_right', type = float, required = False, default = 0.10,  help = 'valid kernel ratio copmpared with ear width from right')
    ap.add_argument('-vkrt', '--valid_kernel_ratio_top', type = float, required = False, default = 0.30,  help = 'valid kernel ratio copmpared with ear length from top')
    ap.add_argument('-vkrb', '--valid_kernel_ratio_bottom', type = float, required = False, default = 0.10,  help = 'valid kernel ratio copmpared with ear length from bottom')
    args = vars(ap.parse_args())
    
    
    # parse input arguments
    file_path = args["path"]
    ext = args['filetype']
    
    coin_path = args["marker"]
    barcode_path = args["barcode"]
    
    min_size = args['min_size']
    min_distance_value = args['min_dist']
    
    coin_size = args['coin_size']
    
    valid_kernel_ratio_list = [0] * 4
    
    valid_kernel_ratio_list[0] = args['valid_kernel_ratio_left']
    valid_kernel_ratio_list[1]  = args['valid_kernel_ratio_right']
    valid_kernel_ratio_list[2] = args['valid_kernel_ratio_top']
    valid_kernel_ratio_list[3]  = args['valid_kernel_ratio_bottom']
    
    
    n_ear = args['num_ears']
    
    # path of the marker (coin), default path will be '/marker_template/marker.png' and '/marker_template/barcode.png'
    # can be changed based on requirement
    global  tp_coin, tp_barcode

    
    #setup marker path to load template
    template_path = file_path + coin_path
    barcode_path = file_path + barcode_path
    
    try:
        # check to see if file is readable
        with open(template_path) as tempFile:

            # Read the template 
            tp_coin = cv2.imread(template_path, 0)
            print("Template loaded successfully...")
            
    except IOError as err:
        
        print("Error reading the Template file {0}: {1}".format(template_path, err))
        exit(0)

    
    try:
        # check to see if file is readable
        with open(barcode_path) as tempFile:

            # Read the template 
            tp_barcode = cv2.imread(barcode_path, 0)
            print("Barcode loaded successfully...\n")
            
    except IOError as err:
        
        print("Error reading the Barcode file {0}: {1}".format(barcode_path, err))
        exit(0)
    
    

    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype
    
    #accquire image file list
    imgList = sorted(glob.glob(image_file_path))
    
    imgList = natsort.natsorted(imgList,reverse = False)
    

    #print((imgList))
    
    n_images = len(imgList)
    
    result_list = []
    
    result_list_cm = []

    ######################################################################################
    #loop execute to get all traits
    for image in imgList:
        
        (image_file_name, tag_info, ear_index_rec, ear_area_rec, ear_area_ratio_rec, ear_width_rec, ear_height_rec, \
            n_kernels_all_rec, kernel_size_rec, n_kernels_valid_rec, kernel_valid_size_rec, \
            diameter_circle, coin_size, pixel_cm_ratio) = extract_traits(image)

       
        for i in range(len(ear_index_rec)):
            
            result_list.append([image_file_name, tag_info, ear_index_rec[i], ear_area_rec[i]/pow(pixel_cm_ratio,2), ear_area_ratio_rec[i], ear_width_rec[i]/pixel_cm_ratio, ear_height_rec[i]/pixel_cm_ratio, \
                                    n_kernels_all_rec[i], kernel_size_rec[i]/pow(pixel_cm_ratio,2), n_kernels_valid_rec[i], kernel_valid_size_rec[i]/pow(pixel_cm_ratio,2), \
                                    diameter_circle/pixel_cm_ratio, coin_size, pixel_cm_ratio])


    #############################################################################################
    #print out result on screen output as table

    print("Summary: {0} plant images were processed...\n".format(n_images))
    
    #output in command window in a sum table
   
    table = tabulate(result_list, headers = ['filename', 'tag_info', 'ear_index', 'ear_area', 'ear_area_ratio', 'ear_width', 'ear_height', 'number_kernels_all', 'kernel_size', 'number_kernels_valid', 'kernel_size_valid', 'coins_diameter', 'coin_size', 'pixel_cm_ratio'], tablefmt = 'orgtbl')
    
    print(table + "\n")
    
    '''
    ####################################################################################
    # parallel processing
    
    # get cpu number for parallel processing
    agents = psutil.cpu_count() - 2
    #agents = multiprocessing.cpu_count() 
    #agents = 8
    
    print("Using {0} cores to perfrom parallel processing... \n".format(int(agents)))
    
    # Create a pool of processes. By default, one is created for each CPU in the machine.
    # extract the bouding box for each image in file list
    with closing(Pool(processes = agents)) as pool:
        result = pool.map(extract_traits, imgList)
        pool.terminate()
    
    
    # unwarp all the computed trait values
    filename = list(zip(*result))[0]
    tag_info = list(zip(*result))[1]
    avg_kernel_size = list(zip(*result))[2]
    avg_n_kernels_valid = list(zip(*result))[3]
    avg_n_kernels_all = list(zip(*result))[4]
    avg_kernel_area = list(zip(*result))[5]
    avg_kernel_area_ratio = list(zip(*result))[6]
    avg_width = list(zip(*result))[7]
    avg_height = list(zip(*result))[8]
    coins_width_avg = list(zip(*result))[9]
    coin_size = list(zip(*result))[10]
    pixel_cm_ratio = list(zip(*result))[11]
    brightness = list(zip(*result))[12]
    blurry_value = list(zip(*result))[13]
    
    '''

    '''
    # create result list
    for i, (v0,v1,v2,v3,v4,v5,v6,v7,v8,v9,v10,v11,v12,v13) in enumerate(zip(filename, tag_info, avg_kernel_size, avg_n_kernels_valid, avg_n_kernels_all, avg_kernel_area, avg_kernel_area_ratio, avg_width, avg_height, coins_width_avg, coin_size, pixel_cm_ratio, brightness, blurry_value)):

        result_list.append([v0,v1,v2,v3,v4,v5,v6,v7,v8,v9,v10,v11,v12,v13])
        
        result_list_cm.append([v0,v1,v2/pow(v11,2),v3,v4,v5/pow(v11,2),v6,v7/v11,v8/v11,v9/v11,v10,v11,v12,v13])
    
    


    '''
    ##############################################################################################
    # save computation traits results as excel file
    
    if (args['result']):

        trait_file = (args['result'] + 'trait.xlsx')

    else:
        trait_file = (file_path + 'trait.xlsx')

    
    # if excel file exits, clear sheets content
    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        #sheet = wb['trait_pixel']

        #sheet_pixel.delete_rows(2, sheet_pixel.max_row - 1) # for entire sheet

        #Get the current Active Sheet
        sheet_cm = wb['traits']

        sheet_cm.delete_rows(2, sheet_cm.max_row - 1) # for entire sheet
        
    # if excel file does not exit, create sheets content
    else:
        # Keep presets
        wb = openpyxl.Workbook()
        
        #sheet = wb.active
        
        sheet_cm = wb.active
        sheet_cm.title = "traits"
    
    
    if pixel_cm_ratio == 1:
        unit_area = '_pixel\u00b2'
        unit_length = '_pixel'
    else:
        unit_area = '_cm\u00b2'
        unit_length = '_cm'
    
            
    # assign traits label names
    sheet_cm.cell(row = 1, column = 1).value = 'filename'
    sheet_cm.cell(row = 1, column = 2).value = 'tag_info'
    sheet_cm.cell(row = 1, column = 3).value = 'ear_index'
    sheet_cm.cell(row = 1, column = 4).value = 'ear_area' + unit_area
    sheet_cm.cell(row = 1, column = 5).value = 'ear_area_ratio' 
    sheet_cm.cell(row = 1, column = 6).value = 'ear_width' + unit_length
    sheet_cm.cell(row = 1, column = 7).value = 'ear_height' + unit_length
    sheet_cm.cell(row = 1, column = 8).value = 'number_kernels_all'
    sheet_cm.cell(row = 1, column = 9).value = 'kernel_size' + unit_area
    sheet_cm.cell(row = 1, column = 10).value = 'number_kernels_valid'
    sheet_cm.cell(row = 1, column = 11).value = 'kernel_size_valid' + unit_area
    sheet_cm.cell(row = 1, column = 12).value = 'coins_diameter' + unit_length
    sheet_cm.cell(row = 1, column = 13).value = 'coin_size' + unit_length
    sheet_cm.cell(row = 1, column = 14).value = 'pixel/cm_ratio'
        
    
    for row in result_list:
        sheet_cm.append(row)
        
   
    
    #save computation traits results as xlsx format excel file
    wb.save(trait_file)
    
    
    


    

    
